"""Writes the e2e results workbook. Driven by scripts/e2eReport.mjs, which
hands it a JSON payload path; keeping the marshalling in JS and only the
spreadsheet here means the run log stays the source of truth."""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

payload = json.load(open(sys.argv[1], encoding="utf-8"))

HEAD = PatternFill("solid", fgColor="1F3A5F")
HEAD_FONT = Font(color="FFFFFF", bold=True)
GREEN = PatternFill("solid", fgColor="D7F0DC")
RED = PatternFill("solid", fgColor="FBD8D8")
AMBER = PatternFill("solid", fgColor="FDEBC8")
GREY = PatternFill("solid", fgColor="EDEDED")


def sheet(wb, title, headers, rows, widths=None, freeze="A2"):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD
        cell.font = HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for r in rows:
        ws.append(r)
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for i, h in enumerate(headers, start=1):
        w = (widths or {}).get(h, min(max(14, len(str(h)) + 4), 46))
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws


wb = Workbook()
wb.remove(wb.active)

# ── Summary ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Summary")
ws.append(["Form 5472 portal — browser test run"])
ws["A1"].font = Font(size=14, bold=True)
ws.append([])
ws.append(["Metric", "Value"])
for c in ("A3", "B3"):
    ws[c].fill = HEAD
    ws[c].font = HEAD_FONT
for k, v in payload["summary"]:
    ws.append([k, v])
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 14

# ── Results ───────────────────────────────────────────────────────────────
rows = payload["rows"]
if rows:
    headers = list(rows[0].keys())
    labels = {
        "scenario_id": "#", "title": "Scenario", "branch": "Branch under test",
        "kind": "Kind", "signature_mode": "Signature", "expected": "Expected",
        "outcome": "Outcome", "stage_reached": "Stage reached",
        "error_text": "Error shown", "console_errors": "Console errors",
        "filing_id": "Filing id", "pdf_count": "PDFs", "pdf_files": "PDF files",
        "pdf_bytes": "PDF bytes", "attempts": "Attempts", "earlier_attempt": "Earlier attempt",
        "signature_checked": "Signature checked",
        "signature_note": "Signature note", "field_mismatches": "Mismatches",
        "notes": "Notes",
    }
    ws = sheet(
        wb, "Results",
        [labels.get(h, h.replace("in_", "entered: ").replace("_", " ")) for h in headers],
        [[r[h] for h in headers] for r in rows],
        widths={"#": 6, "Scenario": 46, "Branch under test": 50, "Expected": 44,
                "Outcome": 26, "Error shown": 44, "Notes": 40},
    )
    oc = headers.index("outcome") + 1
    for row in range(2, len(rows) + 2):
        v = str(ws.cell(row=row, column=oc).value)
        fill = (GREEN if v.startswith("PASS") else
                RED if v.startswith("FAIL") else
                AMBER if v.startswith(("BLOCKED", "ERROR")) else GREY)
        ws.cell(row=row, column=oc).fill = fill

# ── Field check ───────────────────────────────────────────────────────────
fr = payload["fieldRows"]
ws = sheet(
    wb, "Field check",
    ["#", "Scenario", "Field", "Entered", "Saved / on the PDF", "Source", "Severity"],
    [[f["scenario_id"], f["title"], f["field"], f["entered"], f["saved"], f["source"], f["severity"]]
     for f in fr],
    widths={"#": 6, "Scenario": 42, "Field": 30, "Entered": 32,
            "Saved / on the PDF": 32, "Source": 18, "Severity": 12},
)
for row in range(2, len(fr) + 2):
    sev = str(ws.cell(row=row, column=7).value)
    ws.cell(row=row, column=7).fill = GREEN if sev == "match" else RED

# ── Inputs ────────────────────────────────────────────────────────────────
# Every value that went in, one row per field. Banded by scenario so a block
# can be read as one filing rather than as a flat list.
ir = payload["inputRows"]
ws = sheet(
    wb, "Inputs (all fields)",
    ["#", "Scenario", "Section", "Field", "Value entered"],
    ir,
    widths={"#": 6, "Scenario": 44, "Section": 18, "Field": 38, "Value entered": 46},
)
for row in range(2, len(ir) + 2):
    if ir[row - 2][0] % 2 == 0:
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = GREY

# ── PDFs ──────────────────────────────────────────────────────────────────
sheet(
    wb, "PDFs",
    ["#", "File the browser produced", "Bytes"],
    [[p["scenario_id"], p["file"], p["bytes"]] for p in payload["pdfs"]],
    widths={"#": 6, "File the browser produced": 60, "Bytes": 14},
)

wb.save(payload["out"])
print(f"wrote {payload['out']}")
